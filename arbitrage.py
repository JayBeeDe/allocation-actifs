#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import multiprocessing
import openpyxl
import textwrap
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from pylogger_unified import logger as pylogger_unified
import constants
import utils

args = utils.parse_args()
utils.check_args(args)

logger = pylogger_unified.init_logger(json_formatter=False, enable_gi=False, debug=args.debug, logger_name="arbitrage")


def check_fees(data):
    # this function takes fund data from API call and a list of keys of well known fees
    # if a fund has a fee that is not part of the list it will raise an error
    # this avoids missing fees on a fund, impacting decision

    fees = data["fees"]["fees_timed"]

    for fee_key, fee in fees.items():
        if fee_key in constants.known_fee_keys:
            continue
        if not isinstance(fee, dict):
            continue
        if "value" not in fee:
            continue
        if fee["value"] == None:
            continue
        if isinstance(fee["value"], str) and (fee["value"] == "" or fee["value"] == "0"):
            continue
        if isinstance(fee["value"], int) and fee["value"] == 0:
            continue
        logger.error(f"{fee_key} is unknown and its value is {fee['value']}")
        return False

    return True


def get_publication_url(language, publications, doc_type="DOC_KID_PRIIPS"):
    # Document d'informations clés
    urls = []
    if language in publications:
        urls = [d["url"] for d in publications[language]["documents"] if d.get("doc_type") == doc_type]
    if not len(urls):
        return None
    return urls[0]


def get_fund_data(fund):
    global args
    output_item = dict()

    logger = pylogger_unified.init_logger(json_formatter=False, enable_gi=False, debug=args.debug, logger_name=fund)
    logger.info("Getting fund...")

    api_response = utils.request_data(
        url=f"{constants.api_endpoint}/push/fundsheet/{constants.type_to_api_prefix[args.type]}/{args.language}/{args.country}/{fund.lower()}"
    )

    if not api_response:
        logger.error("Failed to retrieve data from the API")

    # Access specific values from the dictionary:
    try:
        ### PRESENTATION ###

        output_item["asset_class"] = api_response["classification"]["asset_class"]
        # Classe d'actif

        output_item["asset_region_class"] = api_response["classification"]["region_reporting"]
        # Région de diversification

        output_item["fundshare_id"] = int(api_response["fundshare_id"])
        # ID du fond

        output_item["legal_name"] = api_response["legal_name"]
        # Nom légal

        output_item["legal_form"] = api_response["portfolio"]["legal_form"]
        # Forme juridique

        output_item["creation_date"] = api_response["portfolio"]["creation_date"]
        # Date de création

        output_item["share_type"] = api_response["fundshare_selection"]["share_types"][str(output_item["fundshare_id"])]
        # Type de parts

        base_currency_code_key = api_response["portfolio"]["base_currency_code"]
        if "share_size" not in api_response["nav"]["nav_info"][base_currency_code_key]:
            base_currency_code_key = "EUR"

        output_item["share_size"] = str(int(api_response["nav"]["nav_info"][base_currency_code_key]["share_size"])) + constants.currency_code_to_symbol[base_currency_code_key]
        # Actif total de la part

        output_item["share_vl"] = str(round(float(api_response["nav"]["two_latest_nav"][base_currency_code_key][0]["nav"]), 2)) + constants.currency_code_to_symbol[base_currency_code_key]
        # Valeur liquidative

        isin = api_response["fundshare_selection"]["share_types_isin_codes"][str(
            output_item["fundshare_id"])]
        # ISIN
        if isin != fund:
            raise ValueError(f"ISIN mismatch (expected {fund}, got {isin})")
        output_item["isin"] = isin

        output_item["currency"] = api_response["portfolio"]["base_currency"]
        if output_item["currency"] == "Euro" and api_response["performances"]["disclaimers"]["currency_fluctuation_not_euro"]["EUR"] and api_response["performances"]["disclaimers"]["currency_fluctuation_not_euro"]["EUR"] is None:
            raise ValueError("Base currency mismatch with currency disclaimer for fund " + fund)
        # Devise

        output_item["base_index"] = api_response["overview"]["bench"]["name"].split(" + ")
        # Indice de référence

        output_item["sri_risk"] = int(api_response["risk"]["sri_risk"]["value"] if api_response["risk"]["sri_risk"]["value"] else 0)
        # Indicateur de risque

        output_item["morning_star"] = int(api_response["fundshare_selection"]["morning_star"]) if api_response["fundshare_selection"]["morning_star"] else 0
        # MorningStar

        output_item["pea"] = "Yes" if bool(api_response["fundshare_selection"]["flags"]["pea_flag"]) else "No"
        # Eligible PEA

        output_item["source_details"] = {
            "url": f"{constants.website_domain}/fr-fr/{constants.type_to_website_prefix[args.type]}/fundsheet/{api_response['fundsheet_uri']}?tab=overview",
            "title": "FR"
        }
        # Source

        output_item["policy"] = textwrap.wrap(api_response["overview"]["disclaimers"]["investment_policy"], width=40)
        # Politique

        ### PERFORMANCES ###

        perf_cumulated = None
        if api_response["performances"]["perfs"] and api_response["performances"]["perfs"]["cumulated"]["shares"]:
            perf_cumulated = next((round(float(perf["value"]), 2) for perf in api_response["performances"]["perfs"]["cumulated"]["shares"] if perf["type"] == "INDEXTYPE_5Y" and perf["currency"] == api_response["portfolio"]["base_currency_code"]), None)
            if perf_cumulated is None:
                logger.warning(f"INDEXTYPE_5Y shares with currency {api_response['portfolio']['base_currency_code']} not found")
        if perf_cumulated is None:
            output_item["perf_cumulated"] = "N/A"
        else:
            output_item["perf_cumulated"] = str(perf_cumulated) + " %"
        # Performance cumulée sur 5 ans

        perf_base_cumulated = None
        if perf_cumulated is not None and api_response["performances"]["perfs"]["cumulated"]["benches"]:
            perf_base_cumulated = next((round(float(perf["value"]), 2) for perf in api_response["performances"]["perfs"]["cumulated"]["benches"] if perf["type"] == "INDEXTYPE_5Y" and perf["currency"] == api_response["portfolio"]["base_currency_code"]), None)
            if perf_base_cumulated is None:
                logger.warning(f"INDEXTYPE_5Y benches with currency {api_response['portfolio']['base_currency_code']} not found")
        if perf_base_cumulated is None:
            output_item["perf_cumulated_diff"] = "N/A"
        else:
            output_item["perf_cumulated_diff"] = str(round(perf_cumulated - perf_base_cumulated, 2)) + " %"
        # Diff indice de base

        try:
            output_item["volatility"] = round(float(api_response["performances"]["risk_analysis"]["stats"]["volatility"]), 2)
            # Volatilité
        except KeyError as e:
            logger.warning("Missing volatility")
            output_item["volatility"] = "Unknown"

        try:
            output_item["sharpe_ratio"] = round(float(api_response["performances"]["risk_analysis"]["stats"]["sharpe_ratio"]), 2)
            # Ratio de Sharpe
        except KeyError as e:
            logger.warning("Missing sharpe_ratio")
            output_item["sharpe_ratio"] = "Unknown"

        version_doc = args.language
        publication_url = get_publication_url(version_doc, api_response["publications"])
        if publication_url is None:
            version_doc = "FRE"
            publication_url = get_publication_url(version_doc, api_response["publications"])
        output_item["dic_details"] = ""
        if publication_url is not None:
            output_item["dic_details"] = {
                "url": publication_url,
                "title": version_doc
            }
        # Document d'informations clés

        res = get_more_details_data(fund)

        output_item["q_notation"] = res["notation"]

        output_item["more_details"] = {
            "url": res["url"],
            "title": "FR"
        }
        # Détails

        ### SCENARIOS ###

        output_item.update(get_scenarios(fund))
        # Rendement de tous les scénarios à 5 ans

        ### FRAIS ###

        output_item["fee_conversion_rate"] = round(float(api_response["fees"]["fees_timed"]["maximum_conversion_rate"]["value"]) if api_response["fees"]["fees_timed"]["maximum_conversion_rate"]["value"] else 0.0, 2)
        # Coûts de conversion

        output_item["fee_ongoing_charges"] = round(float(api_response["fees"]["fees_timed"]["estimated_ongoing_charges"]["value"]) if api_response["fees"]["fees_timed"]["estimated_ongoing_charges"]["value"] else 0.0, 2)
        inter = output_item["fee_ongoing_charges"]
        output_item["fee_ongoing_charges"] += round(float(api_response["fees"]["fees_timed"]["at_launch_ongoing_charges"]["value"]) if api_response["fees"]["fees_timed"]["at_launch_ongoing_charges"]["value"] else 0.0, 2)
        if inter != 0 and (output_item["fee_ongoing_charges"] - inter) != 0:
            raise ValueError("estimated_ongoing_charges and at_launch_ongoing_charges are both non-null values for fund " + fund)
        # Frais courants estimés

        output_item["fee_maximum_subscription"] = round(float(api_response["fees"]["fees_timed"]["total_subscription_fees"]["value"]) if api_response["fees"]["fees_timed"]["total_subscription_fees"]["value"] else 0.0, 2)
        # Frais d'entrée max

        output_item["fee_maximum_redemption"] = round(float(api_response["fees"]["fees_timed"]["total_redemption_fees"]["value"]) if api_response["fees"]["fees_timed"]["total_redemption_fees"]["value"] else 0.0, 2)
        # Frais de sortie max

        output_item["fee_real_ongoing"] = round(float(api_response["fees"]["fees_timed"]["real_ongoing_charges"]["value"]) if api_response["fees"]["fees_timed"]["real_ongoing_charges"]["value"] else 0.0, 2)
        # Frais courants réels

        output_item["fee_redemption_acquired"] = round(float(api_response["fees"]["fees_timed"]["redemption_fixed_fees_acquired"]["value"]) if api_response["fees"]["fees_timed"]["redemption_fixed_fees_acquired"]["value"] else 0.0, 2)
        inter = output_item["fee_redemption_acquired"]
        output_item["fee_redemption_acquired"] += round(float(api_response["fees"]["fees_timed"]["maximum_redemption_fixed_fees_acquired"]["value"]) if api_response["fees"]["fees_timed"]["maximum_redemption_fixed_fees_acquired"]["value"] else 0.0, 2)
        if inter != 0 and (output_item["fee_redemption_acquired"] - inter) != 0:
            raise ValueError("redemption_fixed_fees_acquired and maximum_redemption_fixed_fees_acquired are both non-null values for fund " + fund)
        # Commissions de rachat acquises au fonds

        output_item["fee_maximum_management"] = round(float(api_response["fees"]["fees_timed"]["maximum_management_fees"]["value"]) if api_response["fees"]["fees_timed"]["maximum_management_fees"]["value"] else 0.0, 2)
        # Commission de gestion max

        # perf_benchmark_spread ?

        if not check_fees(api_response):
            raise ValueError(f"Fund {fund} contains unknown fees")

        # output_item["perf_maximum_management"] = float(api_response["fees"]["fees_timed"]["maximum_management_fees"]["value"])
        # Cumulée Fond YTD
        # Cumulée Indice de ref YTD
        # Cumulée Fond Début
        # Cumulée Indice de ref Début

    except KeyError as e:
        logger.error(f"KeyError: Key '{e}' not found in the API response")
        raise

    ### PORTEFEUILLE ###

    api_response = utils.request_data(
        url=f"{constants.api_endpoint}/push/holdings/{args.language}/{str(output_item['fundshare_id'])}"
    )

    if not api_response:
        logger.error("Failed to retrieve data from the API for holding " + str(output_item["fundshare_id"]))

    # Access specific values from the dictionary:
    try:
        if "breakdowns" not in api_response or not api_response["breakdowns"]:
            logger.warning("Missing breakdowns for " + str(output_item["fundshare_id"]))
            api_response["breakdowns"] = []
        for breakdown_item in api_response["breakdowns"]:
            res = [b["label"] + " (" + str(round((b["ptf_value"] if b["ptf_value"] else b["bench_value"]) * 100, 2)) + "%)" for b in sorted(breakdown_item["level_1_breakdowns"], key=lambda x: (x["rank"], -x["ptf_value"] if x["ptf_value"] else -x["bench_value"]))]
            if breakdown_item["labels"]["header"] in [subitem for item in constants.breakdowns_mapping.values() for subitem in item]:
                for breakdown_category in constants.breakdowns_mapping:
                    # breakdown_category=countries...currencies...etc
                    if breakdown_item["labels"]["header"] in constants.breakdowns_mapping[breakdown_category]:
                        if "portfolio_" + breakdown_category in output_item:
                            raise ValueError(f"{utils.join_h(constants.breakdowns_mapping[breakdown_category])} override for {breakdown_category} {str(output_item['fundshare_id'])}")
                        output_item["portfolio_" + breakdown_category] = res
            elif breakdown_item["labels"]["header"] not in constants.breakdowns_exclude:
                logger.warning(f"Unknown portfolio breakdown header {breakdown_item['labels']['header']} for {str(output_item['fundshare_id'])}")

    except KeyError as e:
        logger.error(f"KeyError: Key '{e}' not found in the API response for holding {str(output_item['fundshare_id'])}")
        raise

    return output_item


def gather_data():

    funds = args.isin
    if funds is None:
        # all funds selected
        logger.warning("All funds have been selected...Fetching all funds")
        api_response = utils.request_data(
            url=f"{constants.api_endpoint}/push/fundsearchv2/{constants.type_to_api_prefix[args.type]}/{args.language}?without_has_docs=True&action_column_tool=fundpanorama&with_first_navs=false"
        )

        if not api_response:
            logger.error("Failed to retrieve data from the API")
            exit(1)

        try:
            funds = sorted([fund["codes"]["isin"]
                           for fund in api_response["funds"]])
        except KeyError as e:
            logger.error(f"KeyError: Key '{e}' not found in the API response")
            exit(1)

    with multiprocessing.Pool(processes=multiprocessing.cpu_count()) as pool:
        output_data = pool.map(get_fund_data, funds)

    return output_data


def get_scenarios(fund):
    api_response = utils.request_data(
        url=f"{constants.api_endpoint}/push-raw/all_perf_scenarios?isin={fund.lower()}"
    )

    if not api_response:
        logger.error("Failed to retrieve data from the API")

    try:
        if (len(api_response) == 0):
            raise ValueError("No scenarios found for fund" + fund)
        output_data = {
            "scenario_stressed": str(round(float(api_response[-1]["num02120_portfolio_return_stress_scenario_rhp_or_first_call_dat"] * 100), 2)) + " %",
            "scenario_unfavorable": str(round(float(api_response[-1]["num02030_portfolio_return_unfavourable_scenario_rhp_or_first_ca"] * 100), 2)) + " %",
            "scenario_moderate": str(round(float(api_response[-1]["num02060_portfolio_return_moderate_scenario_rhp_or_first_call_d"] * 100), 2)) + " %",
            "scenario_favorable": str(round(float(api_response[-1]["num02090_portfolio_return_favourable_scenario_rhp_or_first_call"] * 100), 2)) + " %",
        }
    except KeyError as e:
        logger.error(f"KeyError: Key '{e}' not found in the API response")
        raise

    return output_data


def get_more_details_data(fund):
    logger = pylogger_unified.init_logger(json_formatter=False, enable_gi=False, debug=args.debug, logger_name=fund)

    api_response = utils.request_data(
        url=f"{constants.more_details_domain}/Recherche/Data",
        method="POST",
        headers={
            "User-Agent": constants.more_details_user_agent
        },
        cookies={
            "bot_mitigation_cookie": constants.more_details_cookie
        },
        data={
            "columns[0][name]": "ID_Produit",
            "columns[1][name]": "nStarRating",
            "order[0][column]": "0",
            "length": "1",
            "Values.sNomOrISIN": fund
        }
    )

    if not api_response:
        logger.error("Failed to retrieve data from the API")

    try:
        product_id = api_response["data"][0]["ID_Produit"]
    except KeyError as e:
        logger.error(f"KeyError: Key '{e}' not found in the API response")
        raise

    try:
        q_notation = api_response["data"][0]["nStarRating"]
    except KeyError as e:
        logger.error(f"KeyError: Key '{e}' not found in the API response")
        raise

    logger.info(f"Third part website product id is {str(product_id)} with notation {str(q_notation)}")

    return {
        "url": f"{constants.more_details_domain}/Fonds/{str(product_id)}",
        "notation": q_notation
    }


def export_to_file(data):
    logger = pylogger_unified.init_logger(json_formatter=False, enable_gi=False, debug=args.debug, logger_name="export")

    workbook = openpyxl.workbook.Workbook()
    worksheet = workbook.active
    worksheet.title = constants.worksheet["title"]
    worksheet.sheet_properties.tabColor = constants.worksheet["color"]

    header_fill = PatternFill(
        start_color="222222",
        end_color="222222",
        fill_type="solid"
    )
    header_side = Side(
        border_style="thick",
        color="FFFFFF"
    )
    header_border = Border(
        bottom=header_side,
        diagonal=header_side,
        horizontal=header_side,
        left=header_side,
        right=header_side,
        top=header_side,
        vertical=header_side
    )
    standard_fill = PatternFill(
        start_color="333333",
        end_color="333333",
        fill_type="solid"
    )
    standard_side = Side(
        border_style="thin",
        color="FFFFFF"
    )
    standard_border = Border(
        bottom=standard_side,
        diagonal=standard_side,
        horizontal=standard_side,
        left=standard_side,
        right=standard_side,
        top=standard_side,
        vertical=standard_side
    )

    i = 1
    for column_group in constants.column_mapping:
        cell = worksheet.cell(
            row=1,
            column=i,
            value=column_group["name"]
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=18
        )
        cell.border = header_border
        cell.fill = header_fill
        worksheet.merge_cells(f"{get_column_letter(i)}1:{get_column_letter(i + len(column_group['items']) - 1)}1")
        i += len(column_group["items"])
    worksheet.row_dimensions[1].height = 30
    i = 1
    for col_title in [subitem["name"] for item in constants.column_mapping for subitem in item["items"]]:
        cell = worksheet.cell(
            row=2,
            column=i,
            value=col_title
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=10
        )
        cell.border = header_border
        cell.fill = header_fill
        i += 1
    worksheet.row_dimensions[2].height = 30
    worksheet.freeze_panes = "B3"
    worksheet.auto_filter.ref = f"A2:{get_column_letter(worksheet.max_column)}2"

    j = 3
    for row_data in data:
        i = 1
        height_factor = 0
        for col_ref in [subitem["ref"] for item in constants.column_mapping for subitem in item["items"]]:
            cell_height = 1
            if isinstance(row_data.get(col_ref, ""), list):
                cell_height = len(row_data.get(col_ref, ""))
                val = "\n".join(row_data.get(col_ref, ""))
            elif isinstance(row_data.get(col_ref, ""), dict):
                if "url" not in row_data.get(col_ref, ""):
                    raise ValueError("Missing url attribute for dict value " + col_ref)
                val = row_data.get(col_ref, "")["url"]
                if "title" in row_data.get(col_ref, ""):
                    val = row_data.get(col_ref, "")["title"]
            else:
                val = row_data.get(col_ref, "")
            val = utils.remove_invalid_xml_chars(val)

            if re.match(r"^(-\s?)?\d+\.\d+\s?%$", str(val)):
                val = float(re.sub(r"\s?%", "", val))
            elif re.match(r"^(-\s?)?\d+\s?%$", str(val)):
                val = float(re.sub(r"\s?%", "", val))

            cell = worksheet.cell(
                row=j,
                column=i,
                value=val,
            )
            cell.number_format = "@"
            if isinstance(row_data.get(col_ref, ""), dict):
                cell.hyperlink = row_data.get(col_ref, "")["url"]
            cell.alignment = Alignment(
                horizontal="center" if cell_height == 1 else "left",
                vertical="center"
            )
            cell.font = Font(
                size=12,
                color="FFFFFF"
            )
            if i == 1:
                cell.font = Font(
                    bold=True
                )
                cell.border = header_border
                cell.fill = header_fill
            else:
                cell.border = standard_border
                cell.fill = standard_fill
            i += 1
            height_factor = min(max(height_factor, cell_height), 20)

        worksheet.row_dimensions[j].height = height_factor * 16
        j += 1

    i = 1
    for conditional_formatting_item in [subitem["conditional-formatting"] if "conditional-formatting" in subitem else {} for item in constants.column_mapping for subitem in item["items"]]:
        column_letter = openpyxl.utils.get_column_letter(i)
        if "fill-mapping" in conditional_formatting_item:
            for filter_equality, filling_color in conditional_formatting_item["fill-mapping"].items():
                worksheet.conditional_formatting.add(
                    f"{column_letter}1:{column_letter}{j}",
                    FormulaRule(
                        formula=[f"${column_letter}1=\"{filter_equality}\""],
                        fill=PatternFill(
                            start_color=filling_color,
                            end_color=filling_color,
                            fill_type="solid"
                        )
                    )
                )
        if "fill-percentile" in conditional_formatting_item:
            worksheet.conditional_formatting.add(f"{column_letter}1:{column_letter}{j}", ColorScaleRule(
                start_type="percentile",
                start_value=0,
                start_color=conditional_formatting_item["fill-percentile"]["start_color"],
                mid_type="percentile",
                mid_value=50,
                mid_color=conditional_formatting_item["fill-percentile"]["mid_color"],
                end_type="percentile",
                end_value=100,
                end_color=conditional_formatting_item["fill-percentile"]["end_color"]
            ))
        i += 1

    # for i in range(i, 1025):
    #     worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].hidden = True
    #     i += 1

    # for j in range(j, 1048577):
    #     worksheet.row_dimensions[j].hidden = True

    column_fixed_sizes = [subitem["width"] if "width" in subitem else 0 for item in constants.column_mapping for subitem in item["items"]]
    for i, column_cells in enumerate(worksheet.columns):
        if i < len(column_fixed_sizes) and column_fixed_sizes[i]:
            width_value = column_fixed_sizes[i]
        else:
            width_value = 0
            for cell in column_cells:
                for cell_line in str(cell.value).split("\n"):
                    width_value = max(width_value, len(cell_line))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width_value + 4

    worksheet.sheet_view.selection[0].activeCell = "A1"
    worksheet.sheet_view.selection[0].sqref = "A1"
    workbook.save(args.file)
    logger.info(f"excel file {args.file} created successfully!")


if __name__ == "__main__":
    data = gather_data()
    logger.pretty(data)
    export_to_file(data=data)
